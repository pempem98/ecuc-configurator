"""
Complete XLSX Loader - Parses ALL Excel columns.

This loader captures EVERY field from customer Excel files (Rx/Tx sheets),
not just basic CAN data. Provides complete traceability and data preservation.
"""

from typing import Dict, Any, List, Optional
import logging
from pathlib import Path

try:
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    raise ImportError(
        "openpyxl is required for XLSX loading. "
        "Install with: pip install openpyxl"
    )

from .base_loader import BaseLoader, ParserError, ValidationError
from ..model import (
    CompleteXLSXDatabase,
    CompleteXLSXMessage,
    CompleteXLSXSignal,
    MessageDirection,
    RxColumnMapping,
    TxColumnMapping,
    InvalidationPolicy,
    SignalStatus,
)


class CompleteXLSXLoader(BaseLoader[CompleteXLSXDatabase]):
    """
    Complete XLSX Loader - Parses ALL 44 columns from Excel.
    
    Unlike XLSXLoader which only captures basic CAN data, this loader
    preserves EVERY field from the customer Excel format, including:
    - All legacy names and mappings
    - BSW layer data (raw values, types, constraints)
    - Application layer data (scaled values, ports, IDTs)
    - Team responsibilities and status
    - Complete AUTOSAR configuration details
    
    Example:
        >>> loader = CompleteXLSXLoader()
        >>> database = loader.load_complete('CAN_ECM.xlsx')
        >>> 
        >>> # Access ALL fields
        >>> signal = database.messages[0].signals[0]
        >>> print(signal.data_element_name)      # BSW layer
        >>> print(signal.port_name)               # App layer
        >>> print(signal.conversion_function)     # Formula
        >>> print(signal.compu_method_name)       # AUTOSAR
    """
    
    # Sheet names
    RX_SHEET_NAME = 'Rx'
    TX_SHEET_NAME = 'Tx'
    
    # Header rows (0-indexed)
    MAIN_HEADER_ROW = 0   # Row 1: "DBC RX Data", etc.
    SUB_HEADER_ROW = 1    # Row 2: "CAN Message Name", etc.
    DATA_START_ROW = 2    # Row 3: First data row
    
    def __init__(self, logger: Optional[logging.Logger] = None):
        """Initialize complete XLSX loader."""
        super().__init__(logger)
        self._workbook: Optional[openpyxl.Workbook] = None
    
    def load(self, file_path: str) -> Dict[str, Any]:
        """
        Load and parse complete XLSX file.
        
        Args:
            file_path: Path to XLSX file
            
        Returns:
            Dictionary with complete parsed data
            
        Raises:
            FileNotFoundError: If file does not exist
            ParserError: If XLSX parsing fails
        """
        path = self._validate_file_exists(file_path)
        self._validate_file_extension(path, ['.xlsx', '.XLSX'])
        
        try:
            self.logger.info(f"Loading complete XLSX: {path}")
            self._workbook = openpyxl.load_workbook(str(path), data_only=True)
            
            data = {
                'file_path': str(path),
                'rx_messages': [],
                'tx_messages': [],
                'nodes': set(),
            }
            
            # Parse Rx sheet
            if self.RX_SHEET_NAME in self._workbook.sheetnames:
                rx_sheet = self._workbook[self.RX_SHEET_NAME]
                data['rx_messages'] = self._parse_rx_sheet_complete(rx_sheet)
                self.logger.info(f"Parsed {len(data['rx_messages'])} RX messages")
            
            # Parse Tx sheet
            if self.TX_SHEET_NAME in self._workbook.sheetnames:
                tx_sheet = self._workbook[self.TX_SHEET_NAME]
                data['tx_messages'] = self._parse_tx_sheet_complete(tx_sheet)
                self.logger.info(f"Parsed {len(data['tx_messages'])} TX messages")
            
            # Extract unique nodes
            data['nodes'] = sorted(data['nodes'])
            
            return data
            
        except Exception as e:
            raise ParserError(f"Failed to parse XLSX: {e}")
    
    def _parse_rx_sheet_complete(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """
        Parse Rx sheet with ALL 44 columns.
        
        Returns list of message dicts with complete signal data.
        """
        messages = {}
        current_message_name = None
        
        # Iterate from data start row
        for row_idx in range(self.DATA_START_ROW + 1, sheet.max_row + 1):
            # Get message name (may be merged cell)
            msg_name_cell = sheet.cell(row_idx, 1).value
            if msg_name_cell and str(msg_name_cell).strip():
                current_message_name = str(msg_name_cell).strip()
            
            if not current_message_name:
                continue
            
            # Get signal name
            signal_name_cell = sheet.cell(row_idx, 3).value
            if not signal_name_cell or not str(signal_name_cell).strip():
                continue
            
            signal_name = str(signal_name_cell).strip()
            
            # Initialize message if not exists
            if current_message_name not in messages:
                msg_id_str = self._get_cell_value(sheet, row_idx, 2)
                messages[current_message_name] = {
                    'message_name': current_message_name,
                    'message_id': self._parse_message_id(msg_id_str) if msg_id_str else 0,
                    'direction': MessageDirection.RX,
                    'signals': []
                }
            
            # Parse complete signal data (ALL 44 columns)
            signal_data = self._parse_rx_signal_complete(sheet, row_idx)
            messages[current_message_name]['signals'].append(signal_data)
        
        return list(messages.values())
    
    def _parse_rx_signal_complete(self, sheet: Worksheet, row: int) -> Dict[str, Any]:
        """Parse RX signal with ALL 44 columns."""
        
        # Helper to get cell value safely
        def get_val(col_idx: int) -> Optional[str]:
            val = self._get_cell_value(sheet, row, col_idx + 1)  # +1 for 1-based
            return val if val and str(val).strip() not in ['-', ''] else None
        
        # Parse using RxColumnMapping
        data = {
            # Core fields (required)
            'signal_name': get_val(RxColumnMapping.SIGNAL_NAME[2]) or '',
            'signal_size': self._parse_int(get_val(RxColumnMapping.SIGNAL_SIZE[2])) or 1,
            'direction': MessageDirection.RX,
            
            # DBC RX Data section
            'units': get_val(RxColumnMapping.UNITS[2]),
            'signal_group': get_val(RxColumnMapping.SIGNAL_GROUP[2]),
            'has_sna': self._parse_yes_no(get_val(RxColumnMapping.HAS_SNA[2])),
            'periodicity': self._parse_int(get_val(RxColumnMapping.PERIODICITY[2])),
            'timeout': self._parse_int(get_val(RxColumnMapping.TIMEOUT[2])),
            'dbc_comment': get_val(RxColumnMapping.DBC_COMMENT[2]),
            'notes': get_val(RxColumnMapping.NOTES[2]),
            
            # Legacy names (RX-specific)
            'legacy_rx_srd_name': get_val(RxColumnMapping.LEGACY_RX_SRD_NAME[2]),
            'legacy_impl_name': get_val(RxColumnMapping.LEGACY_IMPL_NAME[2]),
            
            # Teams Responsible section
            'start_bit': self._parse_int(get_val(RxColumnMapping.START_BIT[2])),
            'main_function': get_val(RxColumnMapping.MAIN_FUNCTION[2]),
            'consumer_swc': get_val(RxColumnMapping.CONSUMER_SWC[2]),
            'status': self._parse_status(get_val(RxColumnMapping.STATUS[2])),
            
            # Received CAN Signal Raw Data section (BSW)
            'data_element_name': get_val(RxColumnMapping.DATA_ELEMENT_NAME[2]),
            'data_type': get_val(RxColumnMapping.DATA_TYPE[2]),
            'data_struct_element': get_val(RxColumnMapping.DATA_STRUCT_ELEMENT[2]),
            'data_constraint': get_val(RxColumnMapping.DATA_CONSTRAINT[2]),
            'compu_method': get_val(RxColumnMapping.COMPU_METHOD[2]),
            'invalid_value': get_val(RxColumnMapping.INVALID_VALUE[2]),
            'type_reference': get_val(RxColumnMapping.TYPE_REFERENCE[2]),
            'invalidation_policy': self._parse_invalidation_policy(get_val(RxColumnMapping.INVALIDATION_POLICY[2])),
            'initial_value': get_val(RxColumnMapping.INITIAL_VALUE[2]),
            'initial_value_const': get_val(RxColumnMapping.INITIAL_VALUE_CONST[2]),
            'timeout_value': get_val(RxColumnMapping.TIMEOUT_VALUE[2]),
            'conversion_function': get_val(RxColumnMapping.CONVERSION_FUNCTION[2]),
            
            # Provided CAN Signal Scaled Data section (Application)
            'long_name': get_val(RxColumnMapping.LONG_NAME[2]),
            'port_name': get_val(RxColumnMapping.PORT_NAME[2]),
            'data_type_scaled': get_val(RxColumnMapping.DATA_TYPE_SCALED[2]),
            'data_struct_element_scaled': get_val(RxColumnMapping.DATA_STRUCT_ELEMENT_SCALED[2]),
            'struct_element_type_ref': get_val(RxColumnMapping.STRUCT_ELEMENT_TYPE_REF[2]),
            'mapped_idt': get_val(RxColumnMapping.MAPPED_IDT[2]),
            'data_constraint_name': get_val(RxColumnMapping.DATA_CONSTRAINT_NAME[2]),
            'signal_min_value': get_val(RxColumnMapping.SIGNAL_MIN_VALUE[2]),
            'signal_max_value': get_val(RxColumnMapping.SIGNAL_MAX_VALUE[2]),
            'compu_method_name': get_val(RxColumnMapping.COMPU_METHOD_NAME[2]),
            'units_scaled': get_val(RxColumnMapping.UNITS_SCALED[2]),
            'initial_value_scaled': get_val(RxColumnMapping.INITIAL_VALUE_SCALED[2]),
            'invalid_value_scaled': get_val(RxColumnMapping.INVALID_VALUE_SCALED[2]),
            'invalidation_policy_scaled': self._parse_invalidation_policy(get_val(RxColumnMapping.INVALIDATION_POLICY_SCALED[2])),
            'ddf_unit': get_val(RxColumnMapping.DDF_UNIT[2]),
        }
        
        return data
    
    def _parse_tx_sheet_complete(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """
        Parse Tx sheet with ALL 43 columns.
        
        Returns list of message dicts with complete signal data.
        """
        messages = {}
        current_message_name = None
        
        for row_idx in range(self.DATA_START_ROW + 1, sheet.max_row + 1):
            # Get message name
            msg_name_cell = sheet.cell(row_idx, 1).value
            if msg_name_cell and str(msg_name_cell).strip():
                current_message_name = str(msg_name_cell).strip()
            
            if not current_message_name:
                continue
            
            # Get signal name
            signal_name_cell = sheet.cell(row_idx, 3).value
            if not signal_name_cell or not str(signal_name_cell).strip():
                continue
            
            signal_name = str(signal_name_cell).strip()
            
            # Initialize message
            if current_message_name not in messages:
                msg_id_str = self._get_cell_value(sheet, row_idx, 2)
                messages[current_message_name] = {
                    'message_name': current_message_name,
                    'message_id': self._parse_message_id(msg_id_str) if msg_id_str else 0,
                    'direction': MessageDirection.TX,
                    'signals': []
                }
            
            # Parse complete signal
            signal_data = self._parse_tx_signal_complete(sheet, row_idx)
            messages[current_message_name]['signals'].append(signal_data)
        
        return list(messages.values())
    
    def _parse_tx_signal_complete(self, sheet: Worksheet, row: int) -> Dict[str, Any]:
        """Parse TX signal with ALL 43 columns."""
        
        def get_val(col_idx: int) -> Optional[str]:
            val = self._get_cell_value(sheet, row, col_idx + 1)
            return val if val and str(val).strip() not in ['-', ''] else None
        
        data = {
            # Core fields
            'signal_name': get_val(TxColumnMapping.SIGNAL_NAME[2]) or '',
            'signal_size': self._parse_int(get_val(TxColumnMapping.SIGNAL_SIZE[2])) or 1,
            'direction': MessageDirection.TX,
            
            # DBC Tx Data section
            'signal_group': get_val(TxColumnMapping.SIGNAL_GROUP[2]),
            'units': get_val(TxColumnMapping.UNITS[2]),
            'has_sna': self._parse_yes_no(get_val(TxColumnMapping.HAS_SNA[2])),
            'periodicity': self._parse_int(get_val(TxColumnMapping.PERIODICITY[2])),
            'dbc_comment': get_val(TxColumnMapping.DBC_COMMENT[2]),
            'notes': get_val(TxColumnMapping.NOTES[2]),
            
            # Teams Responsible section
            'start_bit': self._parse_int(get_val(TxColumnMapping.START_BIT[2])),
            'main_function': get_val(TxColumnMapping.MAIN_FUNCTION[2]),
            'legacy_srd_dd_name': get_val(TxColumnMapping.LEGACY_SRD_DD_NAME[2]),
            'legacy_tx_accessor_name': get_val(TxColumnMapping.LEGACY_TX_ACCESSOR[2]),
            'pmbd_phase': get_val(TxColumnMapping.PMBD_PHASE[2]),
            'pmbd_coe_producer': get_val(TxColumnMapping.PMBD_COE_PRODUCER[2]),
            'producer_swc': get_val(TxColumnMapping.PRODUCER_SWC[2]),
            'status': self._parse_status(get_val(TxColumnMapping.STATUS[2])),
            
            # Transmitted Raw Signal Data section (BSW)
            'data_element_name': get_val(TxColumnMapping.DATA_ELEMENT_NAME[2]),
            'data_type': get_val(TxColumnMapping.DATA_TYPE[2]),
            'data_struct_element': get_val(TxColumnMapping.DATA_STRUCT_ELEMENT[2]),
            'type_reference': get_val(TxColumnMapping.TYPE_REFERENCE[2]),
            'data_constraint': get_val(TxColumnMapping.DATA_CONSTRAINT[2]),
            'compu_method': get_val(TxColumnMapping.COMPU_METHOD[2]),
            'invalid_value': get_val(TxColumnMapping.INVALID_VALUE[2]),
            'initial_value': get_val(TxColumnMapping.INITIAL_VALUE[2]),
            'initial_value_const': get_val(TxColumnMapping.INITIAL_VALUE_CONST[2]),
            'invalidation_policy': self._parse_invalidation_policy(get_val(TxColumnMapping.INVALIDATION_POLICY[2])),
            'conversion_function': get_val(TxColumnMapping.CONVERSION_FUNCTION[2]),
            
            # Provided CAN Signal Values for TX section (Application)
            'long_name': get_val(TxColumnMapping.LONG_NAME[2]),
            'port_name': get_val(TxColumnMapping.PORT_NAME[2]),
            'data_type_scaled': get_val(TxColumnMapping.DATA_TYPE_SCALED[2]),
            'data_struct_element_scaled': get_val(TxColumnMapping.DATA_STRUCT_ELEMENT_SCALED[2]),
            'struct_element_type_ref': get_val(TxColumnMapping.STRUCT_ELEMENT_TYPE_REF[2]),
            'mapped_idt': get_val(TxColumnMapping.MAPPED_IDT[2]),
            'data_constraint_scaled': get_val(TxColumnMapping.DATA_CONSTRAINT_SCALED[2]),
            'signal_min_value': get_val(TxColumnMapping.SIGNAL_MIN_VALUE[2]),
            'signal_max_value': get_val(TxColumnMapping.SIGNAL_MAX_VALUE[2]),
            'compu_method_scaled': get_val(TxColumnMapping.COMPU_METHOD_SCALED[2]),
            'units_scaled': get_val(TxColumnMapping.UNITS_SCALED[2]),
            'initial_value_scaled': get_val(TxColumnMapping.INITIAL_VALUE_SCALED[2]),
            'invalid_value_scaled': get_val(TxColumnMapping.INVALID_VALUE_SCALED[2]),
            'invalidation_policy_scaled': self._parse_invalidation_policy(get_val(TxColumnMapping.INVALIDATION_POLICY_SCALED[2])),
        }
        
        return data
    
    def validate(self, data: Dict[str, Any]) -> bool:
        """Validate complete parsed data."""
        if not isinstance(data, dict):
            raise ValidationError("Data must be a dictionary")
        
        if 'rx_messages' not in data or 'tx_messages' not in data:
            raise ValidationError("Missing rx_messages or tx_messages")
        
        # Validate messages have required fields
        for msg_list in [data['rx_messages'], data['tx_messages']]:
            for msg in msg_list:
                if 'message_name' not in msg or 'message_id' not in msg:
                    raise ValidationError(f"Message missing required fields: {msg}")
                
                if 'signals' not in msg:
                    raise ValidationError(f"Message missing signals: {msg['message_name']}")
        
        return True
    
    def to_model(self, data: Dict[str, Any]) -> CompleteXLSXDatabase:
        """Convert parsed data to CompleteXLSXDatabase model."""
        messages = []
        
        # Convert RX messages
        for msg_data in data.get('rx_messages', []):
            signals = [CompleteXLSXSignal(**sig_data) for sig_data in msg_data['signals']]
            
            # Auto-detect extended frame
            is_extended = msg_data['message_id'] > 0x7FF
            
            message = CompleteXLSXMessage(
                message_name=msg_data['message_name'],
                message_id=msg_data['message_id'],
                is_extended=is_extended,
                direction=MessageDirection.RX,
                signals=signals,
            )
            messages.append(message)
        
        # Convert TX messages
        for msg_data in data.get('tx_messages', []):
            signals = [CompleteXLSXSignal(**sig_data) for sig_data in msg_data['signals']]
            
            is_extended = msg_data['message_id'] > 0x7FF
            
            message = CompleteXLSXMessage(
                message_name=msg_data['message_name'],
                message_id=msg_data['message_id'],
                is_extended=is_extended,
                direction=MessageDirection.TX,
                signals=signals,
            )
            messages.append(message)
        
        # Create database
        database = CompleteXLSXDatabase(
            name=Path(data['file_path']).stem,
            messages=messages,
            nodes=data.get('nodes', []),
        )
        
        return database
    
    def load_complete(self, file_path: str) -> CompleteXLSXDatabase:
        """
        Load Excel file and return complete database model.
        
        Convenience method that combines load(), validate(), and to_model().
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            CompleteXLSXDatabase with all fields populated
            
        Example:
            >>> loader = CompleteXLSXLoader()
            >>> db = loader.load_complete('CAN_ECM.xlsx')
            >>> print(db.get_statistics())
        """
        data = self.load(file_path)
        self.validate(data)
        return self.to_model(data)
    
    # ==================== Helper Methods ====================
    
    def _get_cell_value(self, sheet: Worksheet, row: int, col: int) -> Optional[str]:
        """Get cell value as string, handling None."""
        cell = sheet.cell(row, col)
        if cell.value is None:
            return None
        return str(cell.value).strip()
    
    def _parse_message_id(self, id_str: Optional[str]) -> int:
        """Parse message ID from various formats."""
        if not id_str:
            return 0
        
        id_str = str(id_str).strip().upper()
        
        # Remove 'h' suffix
        if id_str.endswith('H'):
            id_str = id_str[:-1]
        
        # Try hex first (0x prefix or plain hex)
        try:
            if id_str.startswith('0X'):
                return int(id_str, 16)
            else:
                # Try as hex
                return int(id_str, 16)
        except ValueError:
            pass
        
        # Try decimal
        try:
            return int(id_str)
        except ValueError:
            self.logger.warning(f"Could not parse message ID: {id_str}")
            return 0
    
    def _parse_int(self, value: Optional[str]) -> Optional[int]:
        """Parse integer value."""
        if not value:
            return None
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return None
    
    def _parse_yes_no(self, value: Optional[str]) -> bool:
        """Parse Yes/No to boolean."""
        if not value:
            return False
        value = str(value).strip().upper()
        return value in ['YES', 'Y', 'TRUE', '1']
    
    def _parse_status(self, value: Optional[str]) -> Optional[SignalStatus]:
        """Parse signal status."""
        if not value:
            return None
        
        value = str(value).strip().upper()
        status_map = {
            'NEW': SignalStatus.NEW,
            'MODIFIED': SignalStatus.MODIFIED,
            'UNCHANGED': SignalStatus.UNCHANGED,
            'DELETED': SignalStatus.DELETED,
        }
        
        return status_map.get(value)
    
    def _parse_invalidation_policy(self, value: Optional[str]) -> Optional[InvalidationPolicy]:
        """Parse invalidation policy."""
        if not value:
            return None
        
        value = str(value).strip()
        
        # Map common variations
        policy_map = {
            'Keep': InvalidationPolicy.KEEP,
            'Replace': InvalidationPolicy.REPLACE,
            'Dontinvalidate': InvalidationPolicy.DONT_INVALIDATE,
            'External': InvalidationPolicy.EXTERNAL,
        }
        
        return policy_map.get(value)


__all__ = ['CompleteXLSXLoader']
