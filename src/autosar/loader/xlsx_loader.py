"""
XLSX (Excel) file loader for CAN configuration.

Loads and parses Excel files containing CAN message and signal definitions.
"""

from typing import Dict, Any, List, Optional, Set
import logging
from pathlib import Path

try:
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    raise ImportError(
        "openpyxl is required for XLSX loading. "
        "Install with: pip install openpyxl"
    )

from .base_loader import BaseLoader, ParserError, ValidationError, ConversionError
from ..model import (
    CANDatabase, CANMessage, CANSignal,
    XLSXDatabase, XLSXMessage, XLSXSignal,
    MessageDirection, create_xlsx_database,
    ByteOrder, ValueType, SignalType
)


class XLSXLoader(BaseLoader[CANDatabase]):
    """
    Loader for XLSX (Excel) files containing CAN configuration.
    
    Supports customer-specific Excel format with Rx/Tx sheets.
    """
    
    # Sheet names to process
    RX_SHEET_NAME = 'Rx'
    TX_SHEET_NAME = 'Tx'
    
    # Column mappings for Rx sheet (0-indexed after header row)
    RX_COLUMNS = {
        'message_name': 0,      # CAN Message Name
        'message_id': 1,        # CAN Message ID
        'signal_name': 2,       # CAN Signal Name
        'legacy_srd_name': 3,   # Legacy RX SRD Name
        'legacy_impl_name': 4,  # Legacy Implementation name
        'signal_size': 5,       # Signal size [in bits]
        'units': 6,             # Units
        'signal_group': 7,      # CAN Signal Group Name
        'has_sna': 8,           # Signal has SNA?
        'periodicity': 9,       # Signal Periodicity [ms]
    }
    
    # Column mappings for Tx sheet (0-indexed after header row)
    TX_COLUMNS = {
        'message_name': 0,      # CAN Message Name
        'message_id': 1,        # CAN Message ID
        'signal_name': 2,       # CAN Signal Name
        'signal_group': 3,      # CAN Signal Group Name
        'signal_size': 4,       # Signal Size [in bits]
        'units': 5,             # Units
        'has_sna': 6,           # Signal has SNA?
        'periodicity': 7,       # Signal Periodicity [ms]
        'dbc_comment': 8,       # DBC Comment
        'notes': 9,             # Notes
    }
    
    def __init__(self, logger: Optional[logging.Logger] = None):
        """Initialize XLSX loader."""
        super().__init__(logger)
        self._workbook: Optional[openpyxl.Workbook] = None
    
    def load(self, file_path: str) -> Dict[str, Any]:
        """
        Load and parse XLSX file.
        
        Args:
            file_path: Path to XLSX file
            
        Returns:
            Dictionary with parsed data structure
            
        Raises:
            FileNotFoundError: If file does not exist
            ParserError: If XLSX parsing fails
        """
        # Validate file exists and has correct extension
        path = self._validate_file_exists(file_path)
        self._validate_file_extension(path, ['.xlsx', '.XLSX'])
        
        try:
            # Load workbook
            self.logger.info(f"Parsing XLSX file: {path}")
            self._workbook = openpyxl.load_workbook(str(path), data_only=True)
            
            # Extract data from sheets
            rx_messages = []
            tx_messages = []
            
            if self.RX_SHEET_NAME in self._workbook.sheetnames:
                rx_messages = self._parse_rx_sheet(self._workbook[self.RX_SHEET_NAME])
                self.logger.info(f"Loaded {len(rx_messages)} RX messages")
            else:
                self.logger.warning(f"Sheet '{self.RX_SHEET_NAME}' not found")
            
            if self.TX_SHEET_NAME in self._workbook.sheetnames:
                tx_messages = self._parse_tx_sheet(self._workbook[self.TX_SHEET_NAME])
                self.logger.info(f"Loaded {len(tx_messages)} TX messages")
            else:
                self.logger.warning(f"Sheet '{self.TX_SHEET_NAME}' not found")
            
            # Combine messages
            all_messages = rx_messages + tx_messages
            
            # Extract unique nodes from senders
            nodes = self._extract_nodes(all_messages)
            
            data = {
                'version': '1.0',
                'messages': all_messages,
                'nodes': nodes,
                'file_path': str(path),
                'rx_count': len(rx_messages),
                'tx_count': len(tx_messages),
            }
            
            self.logger.info(
                f"Loaded total {len(all_messages)} messages "
                f"({len(rx_messages)} RX, {len(tx_messages)} TX)"
            )
            
            return data
            
        except Exception as e:
            raise ParserError(f"Failed to parse XLSX file: {e}") from e
        finally:
            if self._workbook:
                self._workbook.close()
    
    def _parse_rx_sheet(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        """
        Parse Rx sheet to extract CAN messages and signals.
        
        Args:
            worksheet: Openpyxl worksheet object
            
        Returns:
            List of message dictionaries
        """
        messages = {}  # message_name -> message_data
        
        # Skip first 2 rows (title and header)
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=3, values_only=True), start=3):
            try:
                # Get message info from first columns
                message_name = row[self.RX_COLUMNS['message_name']]
                message_id = row[self.RX_COLUMNS['message_id']]
                signal_name = row[self.RX_COLUMNS['signal_name']]
                
                # Skip empty rows
                if not signal_name:
                    continue
                
                # If message_name is None, use the previous message name (merged cells)
                if message_name is None:
                    # Find the last valid message name
                    if not messages:
                        self.logger.warning(f"Row {row_idx}: Signal without message name, skipping")
                        continue
                    message_name = list(messages.keys())[-1]
                else:
                    # Create new message if not exists
                    if message_name not in messages:
                        parsed_msg_id = self._parse_message_id(message_id) if message_id else 0
                        # Auto-detect extended frame based on ID value
                        is_extended = parsed_msg_id > 0x7FF
                        messages[message_name] = {
                            'name': message_name,
                            'message_id': parsed_msg_id,
                            'is_extended': is_extended,
                            'dlc': 8,  # Default CAN DLC
                            'cycle_time': self._parse_cycle_time(row[self.RX_COLUMNS['periodicity']]),
                            'senders': [],  # RX messages don't have senders in this format
                            'comment': f"RX Message from XLSX",
                            'signals': [],
                            'direction': 'rx',
                        }
                
                # Create signal
                signal = self._create_signal_from_rx_row(row, row_idx)
                if signal:
                    messages[message_name]['signals'].append(signal)
                    
            except Exception as e:
                self.logger.error(f"Error parsing Rx row {row_idx}: {e}")
                continue
        
        return list(messages.values())
    
    def _parse_tx_sheet(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        """
        Parse Tx sheet to extract CAN messages and signals.
        
        Args:
            worksheet: Openpyxl worksheet object
            
        Returns:
            List of message dictionaries
        """
        messages = {}  # message_name -> message_data
        
        # Skip first 2 rows (title and header)
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=3, values_only=True), start=3):
            try:
                # Get message info from first columns
                message_name = row[self.TX_COLUMNS['message_name']]
                message_id = row[self.TX_COLUMNS['message_id']]
                signal_name = row[self.TX_COLUMNS['signal_name']]
                
                # Skip empty rows
                if not signal_name:
                    continue
                
                # If message_name is None, use the previous message name (merged cells)
                if message_name is None:
                    # Find the last valid message name
                    if not messages:
                        self.logger.warning(f"Row {row_idx}: Signal without message name, skipping")
                        continue
                    message_name = list(messages.keys())[-1]
                else:
                    # Create new message if not exists
                    if message_name not in messages:
                        parsed_msg_id = self._parse_message_id(message_id) if message_id else 0
                        # Auto-detect extended frame based on ID value
                        is_extended = parsed_msg_id > 0x7FF
                        messages[message_name] = {
                            'name': message_name,
                            'message_id': parsed_msg_id,
                            'is_extended': is_extended,
                            'dlc': 8,  # Default CAN DLC
                            'cycle_time': self._parse_cycle_time(row[self.TX_COLUMNS['periodicity']]),
                            'senders': ['ECU'],  # TX messages are sent by ECU
                            'comment': row[self.TX_COLUMNS['dbc_comment']] or "TX Message from XLSX",
                            'signals': [],
                            'direction': 'tx',
                        }
                
                # Create signal
                signal = self._create_signal_from_tx_row(row, row_idx)
                if signal:
                    messages[message_name]['signals'].append(signal)
                    
            except Exception as e:
                self.logger.error(f"Error parsing Tx row {row_idx}: {e}")
                continue
        
        return list(messages.values())
    
    def _create_signal_from_rx_row(self, row: tuple, row_idx: int) -> Optional[Dict[str, Any]]:
        """Create signal dictionary from Rx row data."""
        try:
            signal_name = row[self.RX_COLUMNS['signal_name']]
            signal_size = row[self.RX_COLUMNS['signal_size']]
            
            if not signal_name:
                return None
            
            signal = {
                'name': signal_name,
                'start_bit': 0,  # Not available in XLSX, needs to be calculated
                'length': int(signal_size) if signal_size else 8,
                'byte_order': ByteOrder.LITTLE_ENDIAN,  # Default
                'value_type': ValueType.UNSIGNED,  # Default
                'signal_type': SignalType.STANDARD,
                'factor': 1.0,
                'offset': 0.0,
                'minimum': 0.0,
                'maximum': 0.0,
                'unit': str(row[self.RX_COLUMNS['units']]) if row[self.RX_COLUMNS['units']] else '',
                'receivers': ['ECU'],  # RX signals are received by ECU
                'comment': '',
                'value_table': None,
                # Additional XLSX-specific fields
                'signal_group': str(row[self.RX_COLUMNS['signal_group']]) if row[self.RX_COLUMNS['signal_group']] else '',
                'has_sna': self._parse_yes_no(row[self.RX_COLUMNS['has_sna']]),
                'legacy_srd_name': str(row[self.RX_COLUMNS['legacy_srd_name']]) if row[self.RX_COLUMNS['legacy_srd_name']] else '',
                'legacy_impl_name': str(row[self.RX_COLUMNS['legacy_impl_name']]) if row[self.RX_COLUMNS['legacy_impl_name']] else '',
            }
            
            return signal
            
        except Exception as e:
            self.logger.error(f"Error creating signal from Rx row {row_idx}: {e}")
            return None
    
    def _create_signal_from_tx_row(self, row: tuple, row_idx: int) -> Optional[Dict[str, Any]]:
        """Create signal dictionary from Tx row data."""
        try:
            signal_name = row[self.TX_COLUMNS['signal_name']]
            signal_size = row[self.TX_COLUMNS['signal_size']]
            
            if not signal_name:
                return None
            
            signal = {
                'name': signal_name,
                'start_bit': 0,  # Not available in XLSX, needs to be calculated
                'length': int(signal_size) if signal_size else 8,
                'byte_order': ByteOrder.LITTLE_ENDIAN,  # Default
                'value_type': ValueType.UNSIGNED,  # Default
                'signal_type': SignalType.STANDARD,
                'factor': 1.0,
                'offset': 0.0,
                'minimum': 0.0,
                'maximum': 0.0,
                'unit': str(row[self.TX_COLUMNS['units']]) if row[self.TX_COLUMNS['units']] else '',
                'receivers': [],  # TX signals sent to other ECUs
                'comment': str(row[self.TX_COLUMNS['dbc_comment']]) if row[self.TX_COLUMNS['dbc_comment']] else '',
                'value_table': None,
                # Additional XLSX-specific fields
                'signal_group': str(row[self.TX_COLUMNS['signal_group']]) if row[self.TX_COLUMNS['signal_group']] else '',
                'has_sna': self._parse_yes_no(row[self.TX_COLUMNS['has_sna']]),
                'notes': str(row[self.TX_COLUMNS['notes']]) if row[self.TX_COLUMNS['notes']] else '',
            }
            
            return signal
            
        except Exception as e:
            self.logger.error(f"Error creating signal from Tx row {row_idx}: {e}")
            return None
    
    def _parse_message_id(self, message_id: Any) -> int:
        """
        Parse message ID from various formats.
        
        Supports:
        - Hex string with 'h' suffix: '5B0h', '0x5B0'
        - Decimal integer: 1456
        - Hex string: 'CCh', 'CC'
        """
        if isinstance(message_id, int):
            return message_id
        
        if isinstance(message_id, str):
            message_id = message_id.strip()
            
            # Remove 'h' suffix if present
            if message_id.lower().endswith('h'):
                message_id = message_id[:-1]
            
            # Parse as hex if starts with 0x or looks like hex
            if message_id.startswith('0x') or message_id.startswith('0X'):
                return int(message_id, 16)
            
            # Try parsing as hex first, then decimal
            try:
                return int(message_id, 16)
            except ValueError:
                try:
                    return int(message_id, 10)
                except ValueError:
                    self.logger.warning(f"Could not parse message ID: {message_id}, using 0")
                    return 0
        
        return 0
    
    def _parse_cycle_time(self, periodicity: Any) -> Optional[int]:
        """Parse cycle time/periodicity value."""
        if periodicity is None:
            return None
        
        try:
            return int(periodicity)
        except (ValueError, TypeError):
            return None
    
    def _parse_yes_no(self, value: Any) -> bool:
        """Parse Yes/No value to boolean."""
        if isinstance(value, bool):
            return value
        
        if isinstance(value, str):
            return value.strip().lower() in ['yes', 'y', 'true', '1']
        
        return False
    
    def _extract_nodes(self, messages: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Extract unique nodes from messages."""
        nodes = set()
        
        # Add senders
        for msg in messages:
            for sender in msg.get('senders', []):
                if sender:
                    nodes.add(sender)
            
            # Add receivers from signals
            for signal in msg.get('signals', []):
                for receiver in signal.get('receivers', []):
                    if receiver:
                        nodes.add(receiver)
        
        return [{'name': node, 'comment': ''} for node in sorted(nodes)]
    
    def validate(self, data: Dict[str, Any]) -> bool:
        """
        Validate parsed data structure.
        
        Args:
            data: Parsed data dictionary
            
        Returns:
            True if valid
            
        Raises:
            ValidationError: If validation fails
        """
        try:
            # Check required keys
            required_keys = ['messages', 'nodes', 'version']
            for key in required_keys:
                if key not in data:
                    raise ValidationError(f"Missing required key: {key}")
            
            # Check messages structure
            if not isinstance(data['messages'], list):
                raise ValidationError("messages must be a list")
            
            # Validate each message has required fields
            for idx, msg in enumerate(data['messages']):
                if 'name' not in msg:
                    raise ValidationError(f"Message {idx} missing 'name'")
                if 'message_id' not in msg:
                    raise ValidationError(f"Message {idx} missing 'message_id'")
                if 'signals' not in msg:
                    raise ValidationError(f"Message {idx} missing 'signals'")
            
            self.logger.info("Data validation passed")
            return True
            
        except Exception as e:
            raise ValidationError(f"Validation failed: {e}") from e
    
    def to_xlsx_model(self, data: Dict[str, Any]) -> XLSXDatabase:
        """
        Convert parsed data to XLSXDatabase model (preserves Excel-specific fields).
        
        Args:
            data: Validated data dictionary
            
        Returns:
            XLSXDatabase model instance
            
        Raises:
            ConversionError: If conversion fails
        """
        try:
            database = create_xlsx_database(data)
            self.logger.info(f"Successfully converted to XLSXDatabase model")
            return database
            
        except Exception as e:
            raise ConversionError(f"Failed to convert to XLSX model: {e}") from e
    
    def to_model(self, data: Dict[str, Any]) -> CANDatabase:
        """
        Convert parsed data to CANDatabase model.
        
        Args:
            data: Validated data dictionary
            
        Returns:
            CANDatabase model instance
            
        Raises:
            ConversionError: If conversion fails
        """
        try:
            # Convert messages
            messages = []
            for msg_data in data['messages']:
                # Convert signals
                signals = []
                for sig_data in msg_data['signals']:
                    signal = CANSignal(
                        name=sig_data['name'],
                        short_name=sig_data['name'],
                        start_bit=sig_data['start_bit'],
                        length=sig_data['length'],
                        byte_order=sig_data['byte_order'],
                        value_type=sig_data['value_type'],
                        signal_type=sig_data['signal_type'],
                        factor=sig_data['factor'],
                        offset=sig_data['offset'],
                        minimum=sig_data['minimum'],
                        maximum=sig_data['maximum'],
                        unit=sig_data['unit'],
                        receivers=sig_data['receivers'],
                        comment=sig_data['comment'],
                        value_table=sig_data.get('value_table'),
                    )
                    signals.append(signal)
                
                # Create message
                message = CANMessage(
                    name=msg_data['name'],
                    short_name=msg_data['name'],
                    message_id=msg_data['message_id'],
                    is_extended=msg_data['is_extended'],
                    dlc=msg_data['dlc'],
                    signals=signals,
                    cycle_time=msg_data.get('cycle_time'),
                    senders=msg_data.get('senders', []),
                    comment=msg_data.get('comment', ''),
                )
                messages.append(message)
            
            # Create database
            database = CANDatabase(
                name=Path(data['file_path']).stem,
                short_name=Path(data['file_path']).stem,
                version=data['version'],
                messages=messages,
                nodes=[],  # Nodes are extracted from messages
                comment=f"Loaded from XLSX: {data['file_path']}",
            )
            
            self.logger.info(f"Successfully converted to CANDatabase model")
            return database
            
        except Exception as e:
            raise ConversionError(f"Failed to convert to model: {e}") from e
